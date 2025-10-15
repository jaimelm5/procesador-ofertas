import streamlit as st
import openpyxl
from openpyxl.cell.cell import MergedCell
import random
import re
import io

st.set_page_config(
    page_title="Procesador de Ofertas",
    page_icon="📊",
    layout="wide"
)

# Funciones auxiliares
def obtener_celda_real(ws, celda):
    if isinstance(celda, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if celda.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col)
    return celda

def incrementar_revision(nombre_oferta):
    match = re.search(r'R(\d+)', nombre_oferta)
    if match:
        numero_actual = int(match.group(1))
        numero_nuevo = numero_actual + 1
        nuevo_nombre = re.sub(r'R\d+', f'R{numero_nuevo}', nombre_oferta)
        return nuevo_nombre
    else:
        return nombre_oferta + ' R1'

def es_subtotal(valor):
    if not valor:
        return False
    valor_str = str(valor).strip().lower()
    return 'total' in valor_str and 'oferta' not in valor_str

def aplicar_aumento_aleatorio(valor, porcentaje_min, porcentaje_max):
    if not valor or valor == 0:
        return valor
    try:
        valor_num = float(valor)
        porcentaje = random.uniform(porcentaje_min, porcentaje_max)
        nuevo_valor = valor_num * (1 + porcentaje / 100)
        return round(nuevo_valor, 2)
    except:
        return valor

def procesar_oferta(archivo_bytes, porcentaje_max):
    """Procesa la oferta y devuelve el archivo modificado"""
    PORCENTAJE_MIN = 1
    CELDA_NOMBRE_OFERTA = 'B8'
    
    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes))
    ws = wb.active
    
    nombre_actual = ws[CELDA_NOMBRE_OFERTA].value
    nombre_nuevo = incrementar_revision(str(nombre_actual))
    ws[CELDA_NOMBRE_OFERTA] = nombre_nuevo
    
    filas_subtotales = {}
    filas_partidas = {}
    seccion_actual = None
    
    stats = {
        'precios_modificados': 0,
        'subtotales': 0,
        'unidades_formateadas': 0,
        'nombre_anterior': nombre_actual,
        'nombre_nuevo': nombre_nuevo
    }
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        celda_cantidad = row[0] if len(row) > 0 else None
        celda_uds = row[1] if len(row) > 1 else None
        celda_precio_unitario = row[4] if len(row) > 4 else None
        celda_subtotal = row[5] if len(row) > 5 else None
        
        if not celda_precio_unitario or not celda_subtotal:
            continue
        
        celda_cantidad = obtener_celda_real(ws, celda_cantidad) if celda_cantidad else None
        celda_uds = obtener_celda_real(ws, celda_uds) if celda_uds else None
        celda_precio_unitario = obtener_celda_real(ws, celda_precio_unitario)
        celda_subtotal = obtener_celda_real(ws, celda_subtotal)
        
        valor_cantidad = celda_cantidad.value if celda_cantidad else None
        valor_uds = celda_uds.value if celda_uds else None
        valor_precio_unitario = celda_precio_unitario.value
        valor_subtotal = celda_subtotal.value
        
        if es_subtotal(valor_cantidad):
            stats['subtotales'] += 1
            formato_original = celda_subtotal.number_format
            filas_subtotales[row_idx] = {
                'celda': celda_subtotal,
                'partidas': filas_partidas.get(seccion_actual, []),
                'formato_original': formato_original,
                'nombre': valor_cantidad
            }
            seccion_actual = None
            continue
        
        if valor_precio_unitario and not es_subtotal(valor_cantidad):
            try:
                precio_num = float(valor_precio_unitario)
                
                fill = celda_precio_unitario.fill
                es_celda_azul = False
                if fill and fill.start_color:
                    color = fill.start_color.rgb if hasattr(fill.start_color, 'rgb') else None
                    if color and color not in ['00000000', 'FFFFFFFF', None]:
                        es_celda_azul = True
                
                if not es_celda_azul and precio_num > 0:
                    try:
                        cantidad = float(valor_cantidad) if valor_cantidad else 1
                    except:
                        cantidad = 1
                    
                    if celda_uds:
                        valor_uds_limpio = valor_uds
                        if valor_uds and isinstance(valor_uds, str):
                            valor_uds_limpio = valor_uds.replace('ud', '').replace('Ud', '').replace('UD', '').strip()
                        
                        try:
                            if valor_uds_limpio:
                                cantidad_uds = float(str(valor_uds_limpio).replace(',', '.'))
                            else:
                                cantidad_uds = cantidad
                            
                            celda_uds.value = f"{cantidad_uds:.2f}ud".replace('.', ',')
                            stats['unidades_formateadas'] += 1
                        except:
                            pass
                    
                    precio_nuevo = aplicar_aumento_aleatorio(precio_num, PORCENTAJE_MIN, porcentaje_max)
                    celda_precio_unitario.value = precio_nuevo
                    celda_precio_unitario.number_format = '0.00'
                    
                    subtotal_nuevo = round(cantidad * precio_nuevo, 2)
                    celda_subtotal.value = subtotal_nuevo
                    celda_subtotal.number_format = '0.00'
                    
                    stats['precios_modificados'] += 1
                    
                    if seccion_actual is None:
                        seccion_actual = row_idx
                        filas_partidas[seccion_actual] = []
                    
                    filas_partidas[seccion_actual].append({
                        'subtotal_nuevo': subtotal_nuevo
                    })
            except:
                pass
    
    subtotales_valores = []
    for fila_subtotal, info in filas_subtotales.items():
        partidas = info['partidas']
        if partidas:
            suma_nueva = sum([p['subtotal_nuevo'] for p in partidas])
            info['celda'].value = round(suma_nueva, 2)
            subtotales_valores.append(round(suma_nueva, 2))
            
            if info['formato_original'] and '€' in str(info['formato_original']):
                info['celda'].number_format = '#,##0.00 "€"'
            else:
                info['celda'].number_format = '#,##0.00 "€"'
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        celda_a = row[0] if len(row) > 0 else None
        celda_a = obtener_celda_real(ws, celda_a) if celda_a else None
        
        if celda_a and celda_a.value and 'TOTAL OFERTA' in str(celda_a.value).upper():
            if len(row) > 4:
                celda_total = row[4]
                celda_total = obtener_celda_real(ws, celda_total)
                
                suma_total = sum(subtotales_valores)
                celda_total.value = round(suma_total, 2)
                celda_total.number_format = '#,##0.00 "€"'
                stats['total_oferta'] = suma_total
                break
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    
    return output, stats

# INTERFAZ
st.title("📊 Procesador de Ofertas")
st.markdown("---")

st.sidebar.header("⚙️ Configuración")
st.sidebar.markdown("""
### Instrucciones:
1. Sube tu archivo Excel de oferta
2. Configura el porcentaje máximo de aumento
3. Haz click en "Procesar Oferta"
4. Descarga el archivo procesado
""")

porcentaje_max = st.sidebar.slider(
    "Porcentaje máximo de aumento (%)",
    min_value=1.0,
    max_value=20.0,
    value=5.0,
    step=0.5
)

st.sidebar.info(f"Rango de aumento: 1% - {porcentaje_max}%")

uploaded_file = st.file_uploader(
    "📤 Sube tu archivo de oferta (Excel)",
    type=['xlsx', 'xlsm'],
    help="Selecciona el archivo Excel de la oferta a procesar"
)

if uploaded_file:
    st.success(f"✅ Archivo cargado: {uploaded_file.name}")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.button("🚀 Procesar Oferta", type="primary", use_container_width=True):
            with st.spinner("Procesando oferta..."):
                try:
                    archivo_bytes = uploaded_file.read()
                    output, stats = procesar_oferta(archivo_bytes, porcentaje_max)
                    
                    st.markdown("---")
                    st.success("✅ ¡Oferta procesada correctamente!")
                    
                    col_a, col_b, col_c = st.columns(3)
                    with col_a:
                        st.metric("Precios modificados", stats['precios_modificados'])
                    with col_b:
                        st.metric("Subtotales recalculados", stats['subtotales'])
                    with col_c:
                        st.metric("Unidades formateadas", stats['unidades_formateadas'])
                    
                    st.info(f"**Revisión anterior:** {stats['nombre_anterior']}")
                    st.info(f"**Nueva revisión:** {stats['nombre_nuevo']}")
                    
                    if 'total_oferta' in stats:
                        st.success(f"**Total oferta:** {stats['total_oferta']:,.2f} €".replace(',', 'X').replace('.', ',').replace('X', '.'))
                    
                    nombre_archivo = f"Oferta_{stats['nombre_nuevo'].replace('/', '_').replace(' ', '_')}.xlsx"
                    
                    st.download_button(
                        label="📥 Descargar Oferta Procesada",
                        data=output,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                
                except Exception as e:
                    st.error(f"❌ Error al procesar el archivo: {str(e)}")
                    st.exception(e)
else:
    st.info("👆 Por favor, sube un archivo Excel para comenzar")

st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray;'>
    <small>Procesador de Ofertas v1.0 | Desarrollado con Streamlit</small>
</div>
""", unsafe_allow_html=True)
